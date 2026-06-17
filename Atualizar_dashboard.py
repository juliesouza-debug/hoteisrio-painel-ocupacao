# -*- coding: utf-8 -*-
"""Consolida dados de Hotel (Excel) + STR (AirDNA CSVs) num JSON para o dashboard."""
import openpyxl, csv, json, os, glob, datetime, unicodedata

BASE = os.path.dirname(os.path.abspath(__file__))

# ---- Regiões canônicas (5) + cidade ----
R_BARRA  = "Barra / Recreio / São Conrado"
R_IPA    = "Ipanema / Leblon"
R_LEME   = "Leme / Copacabana"
R_FLABOT = "Flamengo / Botafogo"
R_CENTRO = "Centro"
REGIONS  = [R_BARRA, R_IPA, R_LEME, R_FLABOT, R_CENTRO]
CIDADE   = "Rio de Janeiro (cidade)"

def norm(s):
    if s is None: return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s).encode("ascii","ignore").decode("ascii")
    return s.lower()

def region_from_label(lbl):
    n = norm(lbl)
    if "barra" in n or "recreio" in n or "conrado" in n: return R_BARRA
    if "ipanema" in n or "leblon" in n: return R_IPA
    if "leme" in n or "copacabana" in n: return R_LEME
    if "botafogo" in n or "flamengo" in n or "gloria" in n: return R_FLABOT
    if "centro" in n: return R_CENTRO
    return None

# Mapeia pasta de bairro -> região canônica (para STR)
def region_from_path(path):
    n = norm(path)
    if any(k in n for k in ["barra","recreio","conrado","joa"]): return R_BARRA
    if any(k in n for k in ["ipanema","leblon"]): return R_IPA
    if any(k in n for k in ["leme","copacabana"]): return R_LEME
    if any(k in n for k in ["botafogo","flamengo"]): return R_FLABOT
    # tudo de REGIÃO CENTRO (catumbi, centro, cidade nova, gamboa, gloria, lapa, sta teresa, etc.)
    if "regiao centro" in n: return R_CENTRO
    return None

# ============ HOTEL (Excel) ============
wb = openpyxl.load_workbook(os.path.join(BASE,
    "Pesquisa de ocupaçaõ mes a mes hotéis de 2018 a 2025.xlsx"), data_only=True)
ws = wb.active

hotel = {r: {} for r in REGIONS}      # região -> {"YYYY-MM": occ}
hotel_city = {}                        # cidade -> {"YYYY-MM": occ}

r = 1
maxr = ws.max_row
while r <= maxr:
    label = ws.cell(r,2).value
    if label and "REGI" in norm(label).upper():
        # linha de cabeçalho de meses -> descobre o ANO pelas datas em C..N
        # ano vem da primeira célula datetime
        year = None
        for c in range(3,15):
            v = ws.cell(r,c).value
            if isinstance(v,(datetime.datetime,datetime.date)):
                year = v.year; break
        # linhas de dados começam 2 abaixo, até achar MEDIA
        rr = r+2
        while rr <= maxr:
            lab = ws.cell(rr,2).value
            nlab = norm(lab)
            if not lab:
                rr+=1
                if rr > r+10: break
                continue
            if "media" in nlab:   # fim do bloco
                # também captura média da cidade
                for ci,c in enumerate(range(3,15), start=1):
                    v = ws.cell(rr,c).value
                    if isinstance(v,(int,float)):
                        key = f"{year}-{ci:02d}"
                        hotel_city[key] = round(float(v),2)
                break
            reg = region_from_label(lab)
            if reg and year:
                for ci,c in enumerate(range(3,15), start=1):
                    v = ws.cell(rr,c).value
                    if isinstance(v,(int,float)):
                        key = f"{year}-{ci:02d}"
                        hotel[reg][key] = round(float(v),2)
            rr+=1
        r = rr
    r+=1

# ============ STR (AirDNA CSVs) ============
def read_csv_map(path, valcol=1):
    """Retorna {'YYYY-MM': float} de um CSV Date,Value."""
    out={}
    with open(path, encoding="utf-8-sig") as f:
        rd=csv.reader(f)
        header=next(rd,None)
        for row in rd:
            if len(row)<=valcol or not row[0]: continue
            d=row[0][:7]  # YYYY-MM
            try: out[d]=float(row[valcol])
            except: pass
    return out

# acumula por região: soma ponderada por "Booked Listings"
# str_occ_w[reg][mes] = soma(occ*listings); str_w[reg][mes]=soma(listings)
def acc():
    return {rg:{} for rg in REGIONS}
occ_num, occ_den = acc(), acc()   # ocupação ponderada
adr_num, adr_den = acc(), acc()   # ADR ponderada (peso = listings)
str_listings = acc()             # nº de imóveis com reservas por mês (soma dos bairros)

# encontra todas as pastas de bairro (que tenham occupancy_last_3_years.csv)
occ_files = glob.glob(os.path.join(glob.escape(BASE),"**","occupancy_last_3_years.csv"), recursive=True)
neigh_count = {rg:set() for rg in REGIONS}
for of in occ_files:
    folder = os.path.dirname(of)
    if norm(os.path.basename(folder)) == norm("RIO DE JANEIRO - CIDADE"):
        continue  # cidade tratada à parte
    reg = region_from_path(folder)
    if not reg: continue
    neigh_count[reg].add(os.path.basename(folder))
    occ = read_csv_map(of)
    # demanda (listings) para peso
    dfile = os.path.join(folder,"occupancyDemand_last_3_years.csv")
    listings = read_csv_map(dfile, valcol=2) if os.path.exists(dfile) else {}
    # ADR
    afile = os.path.join(folder,"rateByDailyAverage_last_3_years.csv")
    adr = read_csv_map(afile) if os.path.exists(afile) else {}
    for m,v in occ.items():
        w = listings.get(m, 1) or 1
        occ_num[reg][m]=occ_num[reg].get(m,0)+v*w
        occ_den[reg][m]=occ_den[reg].get(m,0)+w
    for m,v in adr.items():
        w = listings.get(m, 1) or 1
        adr_num[reg][m]=adr_num[reg].get(m,0)+v*w
        adr_den[reg][m]=adr_den[reg].get(m,0)+w
    for m,v in listings.items():
        str_listings[reg][m]=str_listings[reg].get(m,0)+v

str_occ = {rg:{} for rg in REGIONS}
str_adr = {rg:{} for rg in REGIONS}
for rg in REGIONS:
    for m,den in occ_den[rg].items():
        if den: str_occ[rg][m]=round(occ_num[rg][m]/den,2)
    for m,den in adr_den[rg].items():
        if den: str_adr[rg][m]=round(adr_num[rg][m]/den,2)

# Cidade STR: usa pasta oficial RIO DE JANEIRO - CIDADE (occupancy + revpar -> ADR)
city_folder = None
for of in occ_files:
    if norm(os.path.basename(os.path.dirname(of)))==norm("RIO DE JANEIRO - CIDADE"):
        city_folder=os.path.dirname(of); break
str_occ_city, str_adr_city, str_listings_city = {}, {}, {}
if city_folder:
    occ_c = read_csv_map(os.path.join(city_folder,"occupancy_last_3_years.csv"))
    revpar_c = read_csv_map(os.path.join(city_folder,"revpar_last_3_years.csv"))
    str_occ_city = {m:round(v,2) for m,v in occ_c.items()}
    for m,occv in occ_c.items():
        rp = revpar_c.get(m)
        if rp and occv: str_adr_city[m]=round(rp/(occv/100.0),2)
    dcity = os.path.join(city_folder,"occupancyDemand_last_3_years.csv")
    if os.path.exists(dcity):
        str_listings_city = {m:int(v) for m,v in read_csv_map(dcity, valcol=2).items()}
# fallback: se a cidade não tiver demanda, usa a soma dos bairros
if not str_listings_city:
    tmp={}
    for rg in REGIONS:
        for m,v in str_listings[rg].items(): tmp[m]=tmp.get(m,0)+v
    str_listings_city={m:int(v) for m,v in tmp.items()}
str_listings = {rg:{m:int(v) for m,v in str_listings[rg].items()} for rg in REGIONS}

# ---- Inventário de HOTELARIA (snapshot fixo · fonte HotéisRIO) ----
inv_hotel = {
    R_BARRA:  {"emp":44, "uh":8390},
    R_IPA:    {"emp":18, "uh":1372},
    R_LEME:   {"emp":87, "uh":10970},
    R_FLABOT: {"emp":66, "uh":4972},
    R_CENTRO: {"emp":71, "uh":5940},
    CIDADE:   {"emp":286,"uh":31644},
}

# ============ MONTAGEM JSON ============
def revpar(occmap, adrmap):
    out={}
    for m,occv in occmap.items():
        a=adrmap.get(m)
        if a is not None: out[m]=round(a*occv/100.0,2)
    return out

data = {
    "geradoEm": "2026-06-16",
    "regioes": REGIONS,
    "cidade": CIDADE,
    "bairrosPorRegiao": {rg: sorted(neigh_count[rg]) for rg in REGIONS},
    "hotel": {"occ": {**hotel, CIDADE: hotel_city}},
    "str": {
        "occ": {**str_occ, CIDADE: str_occ_city},
        "adr": {**str_adr, CIDADE: str_adr_city},
        "revpar": {**{rg:revpar(str_occ[rg],str_adr[rg]) for rg in REGIONS},
                   CIDADE: revpar(str_occ_city, str_adr_city)},
        "listings": {**str_listings, CIDADE: str_listings_city},
    },
    "inventarioHotel": inv_hotel,
}

with open(os.path.join(BASE,"_data.json"),"w",encoding="utf-8") as f:
    json.dump(data,f,ensure_ascii=False)

# ---- injeta os dados + logo no template e gera o dashboard final ----
import base64
def logo_data_uri():
    cand = os.path.join(BASE, "logo_hoteisrio_branca.png")
    if not os.path.exists(cand):
        return ""
    b = base64.b64encode(open(cand, "rb").read()).decode("ascii")
    return "data:image/png;base64," + b

tpl_path = os.path.join(BASE,"_template.html")
if os.path.exists(tpl_path):
    tpl = open(tpl_path, encoding="utf-8").read()
    html = tpl.replace("__DATA__", json.dumps(data, ensure_ascii=False))
    html = html.replace("__LOGO__", logo_data_uri())
    open(os.path.join(BASE,"Dashboard_Hotel_vs_Temporada.html"),"w",encoding="utf-8").write(html)
    print("Dashboard_Hotel_vs_Temporada.html atualizado.")

# resumo
print("HOTEL occ meses (cidade):", len(hotel_city), "regiões:", {k:len(v) for k,v in hotel.items()})
print("STR occ meses (cidade):", len(str_occ_city), "regiões:", {k:len(v) for k,v in str_occ.items()})
print("STR adr cidade meses:", len(str_adr_city))
print("Bairros por região:", {k:len(v) for k,v in neigh_count.items()})
allm = sorted(set(list(hotel_city)+list(str_occ_city)))
print("Range hotel:", min(hotel_city) if hotel_city else None, "->", max(hotel_city) if hotel_city else None)
print("Range STR:", min(str_occ_city) if str_occ_city else None, "->", max(str_occ_city) if str_occ_city else None)
print("OK -> _data.json")
